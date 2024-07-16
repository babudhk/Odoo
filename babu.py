import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Select the active sheet
ws = wb.active

# Write data to the sheet
ws['A1'] = 'Hello'
ws['B1'] = 'World'
ws['A2'] = 'This'
ws['B2'] = 'is'
ws['C2'] = 'Python'

# Save the workbook to a file
wb.save('D:/example.xlsx')

print("Excel file created and data written successfully.")